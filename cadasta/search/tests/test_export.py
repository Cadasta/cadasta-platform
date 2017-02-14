import csv
import io
import json
import os
import pytest
import shutil

from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.test import TestCase
from openpyxl import load_workbook
from zipfile import ZipFile

from core.models import RandomIDModel
from core.util import random_id
from core.tests.utils.cases import UserTestCase
from core.tests.utils.files import make_dirs  # noqa
from organization.tests.factories import ProjectFactory
from party.tests.factories import PartyFactory, TenureRelationshipFactory
from resources.models import ContentObject
from resources.tests.factories import ResourceFactory
from resources.tests.utils import clear_temp  # noqa
from resources.utils.io import ensure_dirs
from spatial.tests.factories import SpatialUnitFactory
from questionnaires.managers import create_attrs_schema
from questionnaires.tests import attr_schemas
from questionnaires.tests.factories import QuestionnaireFactory
from .fake_results import get_fake_es_api_results
from ..export.base import Exporter
from ..export.all import AllExporter
from ..export.resource import ResourceExporter
from ..export.shape import ShapeExporter
from ..export.xls import XLSExporter


test_dir = os.path.join(settings.MEDIA_ROOT, 'temp')


class BaseTestClass(UserTestCase, TestCase):
    """Base test class that uses the "standard" test project schema."""

    def setUp(self):
        super().setUp()
        self.project = ProjectFactory.create()
        QuestionnaireFactory.create(project=self.project)

        content_type = ContentType.objects.get(
            app_label='spatial', model='spatialunit')
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.location_xform_group,
            content_type=content_type,
        )
        content_type = ContentType.objects.get(
            app_label='party', model='party')
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.default_party_xform_group,
            content_type=content_type,
        )
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.individual_party_xform_group,
            content_type=content_type,
        )
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.group_party_attributes,
            content_type=content_type,
        )
        content_type = ContentType.objects.get(
            app_label='party', model='tenurerelationship')
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.tenure_relationship_xform_group,
            content_type=content_type,
        )


class ExporterTest(BaseTestClass):

    def test_init(self):
        exporter = Exporter(self.project)
        assert exporter.project == self.project

        metadatum = exporter.metadata['location']
        assert metadatum['title'] == 'locations'
        assert metadatum['model_name'] == 'SpatialUnit'
        assert metadatum['content_type'] == ContentType.objects.get(
            app_label='spatial', model='spatialunit')
        assert list(metadatum['schema_attrs'].keys()) == ['DEFAULT']
        attrs = metadatum['schema_attrs']['DEFAULT']
        assert list(attrs.keys()) == ['quality', 'infrastructure']
        assert metadatum['model_attrs'] == ['id', 'geometry.ewkt', 'type']
        assert list(metadatum['attr_columns'].keys()) == [
            'id', 'geometry.ewkt', 'type', 'quality', 'infrastructure']

        metadatum = exporter.metadata['party']
        assert metadatum['title'] == 'parties'
        assert metadatum['model_name'] == 'Party'
        assert metadatum['content_type'] == ContentType.objects.get(
            app_label='party', model='party')
        assert list(metadatum['schema_attrs'].keys()) == ['IN', 'CO', 'GR']
        attrs = metadatum['schema_attrs']['IN']
        assert list(attrs.keys()) == ['notes', 'gender', 'homeowner', 'dob']
        attrs = metadatum['schema_attrs']['CO']
        assert list(attrs.keys()) == ['notes']
        attrs = metadatum['schema_attrs']['GR']
        assert list(attrs.keys()) == [
            'notes', 'number_of_members', 'date_formed']
        assert metadatum['model_attrs'] == ['id', 'name', 'type']
        assert list(metadatum['attr_columns'].keys()) == [
            'id', 'name', 'type', 'notes', 'gender', 'homeowner', 'dob',
            'number_of_members', 'date_formed']

        metadatum = exporter.metadata['tenure_rel']
        assert metadatum['title'] == 'relationships'
        assert metadatum['model_name'] == 'TenureRelationship'
        assert metadatum['content_type'] == ContentType.objects.get(
            app_label='party', model='tenurerelationship')
        assert list(metadatum['schema_attrs'].keys()) == ['DEFAULT']
        attrs = metadatum['schema_attrs']['DEFAULT']
        assert list(attrs.keys()) == ['notes']
        assert metadatum['model_attrs'] == ['id', 'party_id',
                                            'spatial_unit_id',
                                            'tenure_type.id',
                                            'tenure_type.label']
        assert list(metadatum['attr_columns'].keys()) == [
            'id', 'party_id', 'spatial_unit_id', 'tenure_type.id',
            'tenure_type.label', 'notes']

    def test_get_attr_values(self):
        location_data = {
            'id': 'ID',
            'geometry': {'ewkt': 'GEOMETRY.EWKT'},
            'type': 'TYPE',
            'attributes': {
                'quality': 'QUALITY',
                'infrastructure': 'INFRASTRUCTURE',
            },
        }

        exporter = Exporter(self.project)
        metadatum = exporter.metadata['location']
        attr_values = exporter.get_attr_values(location_data, metadatum)
        for key in metadatum['attr_columns'].keys():
            assert attr_values[key] == key.upper()

    def test_get_attr_values_with_conditional_selector(self):
        party_data = {
            'id': 'ID',
            'name': 'NAME',
            'type': 'IN',
            'attributes': {
                'notes': 'NOTES',
                'gender': 'GENDER',
                'homeowner': 'HOMEOWNER',
                'dob': 'DOB',
            },
        }

        exporter = Exporter(self.project)
        metadatum = exporter.metadata['party']
        attr_values = exporter.get_attr_values(party_data, metadatum)
        for key in metadatum['attr_columns'].keys():
            if key == 'type':
                assert attr_values[key] == 'IN'
            elif key in ['number_of_members', 'date_formed']:
                assert key not in attr_values
            else:
                assert attr_values[key] == key.upper()

    def test_get_attr_values_with_list_attr(self):
        tenure_rel_data = {
            'id': 'ID',
            'party_id': 'PARTY_ID',
            'spatial_unit_id': 'SPATIAL_UNIT_ID',
            'tenure_type': {
                'id': 'TENURE_TYPE.ID',
                'label': 'TENURE_TYPE.LABEL',
            },
            'attributes': {
                'notes': ['1', '2', '3'],
            },
        }

        exporter = Exporter(self.project)
        metadatum = exporter.metadata['tenure_rel']
        attr_values = exporter.get_attr_values(tenure_rel_data, metadatum)
        for key in metadatum['attr_columns'].keys():
            if key == 'notes':
                assert attr_values[key] == '1, 2, 3'
            else:
                assert attr_values[key] == key.upper()

    def test_process_location_entity(self):
        es_type_line = '{"index": {"_type": "spatial"} }'
        dummies = []
        for _ in range(5):
            obj = RandomIDModel()
            obj.id = random_id()
            dummies.append(obj)
        es_source_line = json.dumps(
            get_fake_es_api_results(*dummies)['hits']['hits'][1]['_source'])

        exporter = Exporter(self.project)

        def callback(source, metadatum):
            assert metadatum == exporter.metadata['location']
            assert source['id'] == dummies[1].id
            assert source['geometry']['ewkt'] == ('SRID=4326;POINT (1 1)')
            assert source['geometry']['wkt'] == ('POINT (1 1)')
            assert source['attributes']['name'] == "Long Island"
            assert source['attributes']['acquired_how'] == 'LH'

        exporter.process_entity(es_type_line, es_source_line, callback)

    def test_process_party_entity(self):
        es_type_line = '{"index": {"_type": "party"} }'
        dummies = []
        for _ in range(5):
            obj = RandomIDModel()
            obj.id = random_id()
            dummies.append(obj)
        es_source_line = json.dumps(
            get_fake_es_api_results(*dummies)['hits']['hits'][2]['_source'])

        exporter = Exporter(self.project)

        def callback(source, metadatum):
            assert metadatum == exporter.metadata['party']
            assert source['id'] == dummies[2].id
            assert source['name'] == "Party in the USA"
            assert source['attributes']['party_notes'] == "PBS is the best."

        exporter.process_entity(es_type_line, es_source_line, callback)

    def test_process_tenure_rel_entity(self):
        es_type_line = '{"index": {"_type": "party"} }'
        dummies = []
        for _ in range(5):
            obj = RandomIDModel()
            obj.id = random_id()
            dummies.append(obj)
        es_source_line = json.dumps(
            get_fake_es_api_results(*dummies)['hits']['hits'][3]['_source'])

        exporter = Exporter(self.project)

        def callback(source, metadatum):
            assert metadatum == exporter.metadata['tenure_rel']
            assert source['id'] == dummies[3].id
            assert source['spatial_unit_id'] == dummies[1].id
            assert source['party_id'] == dummies[2].id
            assert source['tenure_type']['id'] == 'CU'
            assert source['tenure_type']['label'] == 'Customary Rights'
            assert source['attributes']['rel_notes'] == "PBS is the best."

        exporter.process_entity(es_type_line, es_source_line, callback)

    def test_process_invalid_entity(self):
        es_type_line = '{"index": {"_type": "project"} }'
        es_source_line = 'DUMMY'

        def callback(source, metadatum):
            assert False  # callback should not be called

        exporter = Exporter(self.project)
        exporter.process_entity(es_type_line, es_source_line, callback)


@pytest.mark.usefixtures('clear_temp')
class ShapeExporterTest(BaseTestClass):

    def test_make_download_standalone(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_basic.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test1.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = ShapeExporter(self.project)
        zip_path, mime_type = exporter.make_download(es_dump_path)
        assert zip_path == os.path.join(test_dir, 'test1-shp.zip')
        assert mime_type == ('application/zip')

        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 8
            assert 'README.txt' in files
            assert 'locations.csv' in files
            assert 'parties.csv' in files
            assert 'relationships.csv' in files
            assert 'point.dbf' in files
            assert 'point.prj' in files
            assert 'point.shp' in files
            assert 'point.shx' in files

            with myzip.open('locations.csv') as csv_file:
                rows = list(csv.reader(io.TextIOWrapper(csv_file)))
                assert rows[0][0] == 'id'
                assert rows[0][1] == 'type'
                assert rows[0][2] == 'quality'
                assert rows[0][3] == 'infrastructure'
                assert rows[1][0] == 'ID0'
                assert rows[1][1] == 'PA'
                assert rows[1][2] == 'point'
                assert rows[1][3] == 'food, electricity'

            with myzip.open('parties.csv') as csv_file:
                rows = list(csv.reader(io.TextIOWrapper(csv_file)))
                assert rows[0][0] == 'id'
                assert rows[0][1] == 'name'
                assert rows[0][2] == 'type'
                assert rows[0][3] == 'notes'
                assert rows[0][4] == 'gender'
                assert rows[0][5] == 'homeowner'
                assert rows[0][6] == 'dob'
                assert rows[0][7] == 'number_of_members'
                assert rows[0][8] == 'date_formed'
                assert rows[1][0] == 'ID1'
                assert rows[1][1] == "Cadastanaut"
                assert rows[1][2] == 'IN'
                assert rows[1][3] is ''
                assert rows[1][4] == 'm'
                assert rows[1][5] == 'yes'
                assert rows[1][6] == '1951-05-05'
                assert rows[1][7] is ''
                assert rows[1][8] is ''

            with myzip.open('relationships.csv') as csv_file:
                rows = list(csv.reader(io.TextIOWrapper(csv_file)))
                assert rows[0][0] == 'id'
                assert rows[0][1] == 'party_id'
                assert rows[0][2] == 'spatial_unit_id'
                assert rows[0][3] == 'tenure_type.id'
                assert rows[0][4] == 'tenure_type.label'
                assert rows[0][5] == 'notes'
                assert rows[1][0] == 'ID2'
                assert rows[1][1] == 'ID1'
                assert rows[1][2] == 'ID0'
                assert rows[1][3] == 'CU'
                assert rows[1][4] == "Customary Rights"
                assert rows[1][5] == "The best relationship!"

    def test_make_download_not_standalone(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_basic.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test2.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = ShapeExporter(self.project, is_standalone=False)
        dir_path = exporter.make_download(es_dump_path)
        assert dir_path == os.path.join(test_dir, 'test2-shp-dir')

        files = os.listdir(dir_path)
        assert len(files) == 5
        assert 'README.txt' in files
        assert 'point.dbf' in files
        assert 'point.prj' in files
        assert 'point.shp' in files
        assert 'point.shx' in files

    def test_make_download_empty_standalone(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_empty.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test-empty1.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = ShapeExporter(self.project, True)
        zip_path, mime_type = exporter.make_download(es_dump_path)

        assert zip_path == os.path.join(test_dir, 'test-empty1-shp.zip')
        assert mime_type == ('application/zip')
        assert ZipFile(zip_path).namelist() == []

    def test_make_download_empty_not_standalone(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_empty.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test-empty2.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = ShapeExporter(self.project, is_standalone=False)
        dir_path = exporter.make_download(es_dump_path)

        assert dir_path == os.path.join(test_dir, 'test-empty2-shp-dir')
        assert os.listdir(dir_path) == []


@pytest.mark.usefixtures('clear_temp')
class XLSExporterTest(BaseTestClass):

    def test_make_download(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_basic.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = XLSExporter(self.project)
        xls_path, mime_type = exporter.make_download(es_dump_path)
        assert xls_path == os.path.join(test_dir, 'test.xlsx')
        assert mime_type == ('application/vnd.openxmlformats-officedocument.'
                             'spreadsheetml.sheet')

        wb = load_workbook(xls_path)
        assert wb.get_sheet_names() == [
            'locations', 'parties', 'relationships']

        ws = wb['locations']
        assert ws['A1'].value == 'id'
        assert ws['B1'].value == 'geometry.ewkt'
        assert ws['C1'].value == 'type'
        assert ws['D1'].value == 'quality'
        assert ws['E1'].value == 'infrastructure'
        assert ws['A2'].value == 'ID0'
        assert ws['B2'].value == 'SRID=4326;POINT (1 1)'
        assert ws['C2'].value == 'PA'
        assert ws['D2'].value == 'point'
        assert ws['E2'].value == 'food, electricity'

        ws = wb['parties']
        assert ws['A1'].value == 'id'
        assert ws['B1'].value == 'name'
        assert ws['C1'].value == 'type'
        assert ws['D1'].value == 'notes'
        assert ws['E1'].value == 'gender'
        assert ws['F1'].value == 'homeowner'
        assert ws['G1'].value == 'dob'
        assert ws['H1'].value == 'number_of_members'
        assert ws['I1'].value == 'date_formed'
        assert ws['A2'].value == 'ID1'
        assert ws['B2'].value == "Cadastanaut"
        assert ws['C2'].value == 'IN'
        assert ws['D2'].value is None
        assert ws['E2'].value == 'm'
        assert ws['F2'].value == 'yes'
        assert ws['G2'].value == '1951-05-05'
        assert ws['H2'].value is None
        assert ws['I2'].value is None

        ws = wb['relationships']
        assert ws['A1'].value == 'id'
        assert ws['B1'].value == 'party_id'
        assert ws['C1'].value == 'spatial_unit_id'
        assert ws['D1'].value == 'tenure_type.id'
        assert ws['E1'].value == 'tenure_type.label'
        assert ws['F1'].value == 'notes'
        assert ws['A2'].value == 'ID2'
        assert ws['B2'].value == 'ID1'
        assert ws['C2'].value == 'ID0'
        assert ws['D2'].value == 'CU'
        assert ws['E2'].value == "Customary Rights"
        assert ws['F2'].value == "The best relationship!"

    def test_make_download_empty(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_empty.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = XLSExporter(self.project)
        xls_path, mime_type = exporter.make_download(es_dump_path)
        assert xls_path == os.path.join(test_dir, 'test.xlsx')
        assert mime_type == ('application/vnd.openxmlformats-officedocument.'
                             'spreadsheetml.sheet')

        wb = load_workbook(xls_path)
        assert wb.get_sheet_names() == ['Sheet']
        assert wb['Sheet']['A1'].value is None


@pytest.mark.usefixtures('clear_temp')
@pytest.mark.usefixtures('make_dirs')
class ResourceExporterTest(BaseTestClass):

    def test_make_download(self):
        res = ResourceFactory.create(project=self.project)
        loc1 = SpatialUnitFactory.create(project=self.project)
        loc2 = SpatialUnitFactory.create(project=self.project)
        par = PartyFactory.create(project=self.project)
        rel = TenureRelationshipFactory.create(project=self.project)

        ContentObject.objects.create(resource=res, content_object=loc1)
        ContentObject.objects.create(resource=res, content_object=loc2)
        ContentObject.objects.create(resource=res, content_object=par)
        ContentObject.objects.create(resource=res, content_object=rel)

        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_basic.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test3-orig.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        # Rewrite the ES dump file to inject the resource ID
        with open(es_dump_path, 'r') as infile:
            es_dump_path = os.path.join(test_dir, 'test3.esjson')
            fwrite = open(es_dump_path, 'w')
            for line in infile:
                fwrite.write(line.replace('ID3', res.id))
            fwrite.close()

        exporter = ResourceExporter(self.project)
        zip_path, mime_type = exporter.make_download(es_dump_path)

        assert zip_path == os.path.join(test_dir, 'test3-res.zip')
        assert mime_type == ('application/zip')

        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 2
            assert 'resources.xlsx' in files
            assert 'resources/baby_goat.jpeg' in files

            myzip.extract('resources.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'resources.xlsx'))
            assert wb.get_sheet_names() == ['resources']

            ws = wb['resources']
            assert ws['A1'].value == 'id'
            assert ws['B1'].value == 'name'
            assert ws['C1'].value == 'description'
            assert ws['D1'].value == 'filename'
            assert ws['E1'].value == 'locations'
            assert ws['F1'].value == 'parties'
            assert ws['G1'].value == 'relationships'
            assert ws['A2'].value == res.id
            assert ws['B2'].value == "Goat"
            assert ws['C2'].value == "Let's pretend there's a description."
            assert ws['D2'].value == "baby_goat.jpeg"
            ids = ws['E2'].value.split(',')
            assert len(ids) == 2
            assert loc1.id in ids
            assert loc2.id in ids
            assert ws['F2'].value == par.id
            assert ws['G2'].value == rel.id

    def test_make_download_with_dupe_filenames(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_dupe_resources.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test4.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = ResourceExporter(self.project)
        zip_path, mime_type = exporter.make_download(es_dump_path)

        assert zip_path == os.path.join(test_dir, 'test4-res.zip')
        assert mime_type == ('application/zip')

        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 3
            assert 'resources.xlsx' in files
            assert 'resources/text.csv' in files
            assert 'resources/text (2).csv' in files

            myzip.extract('resources.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'resources.xlsx'))
            assert wb.get_sheet_names() == ['resources']

            ws = wb['resources']
            assert ws['A1'].value == 'id'
            assert ws['B1'].value == 'name'
            assert ws['C1'].value == 'description'
            assert ws['D1'].value == 'filename'
            assert ws['E1'].value == 'locations'
            assert ws['F1'].value == 'parties'
            assert ws['G1'].value == 'relationships'
            assert ws['A2'].value == 'ID3'
            assert ws['B2'].value == "File 1"
            assert ws['C2'].value == "Description 1"
            assert ws['D2'].value == "text.csv"
            assert ws['E2'].value is None
            assert ws['F2'].value is None
            assert ws['G2'].value is None
            assert ws['A3'].value == 'ID4'
            assert ws['B3'].value == "File 2"
            assert ws['C3'].value == "Description 2"
            assert ws['D3'].value == "text (2).csv"
            assert ws['E3'].value is None
            assert ws['F3'].value is None
            assert ws['G3'].value is None

    def test_make_download_empty(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_empty.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test5.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = ResourceExporter(self.project)
        zip_path, mime_type = exporter.make_download(es_dump_path)

        assert zip_path == os.path.join(test_dir, 'test5-res.zip')
        assert mime_type == ('application/zip')
        assert ZipFile(zip_path).namelist() == []


@pytest.mark.usefixtures('clear_temp')
@pytest.mark.usefixtures('make_dirs')
class AllExporterTest(BaseTestClass):

    def test_make_download(self):
        res = ResourceFactory.create(project=self.project)
        loc1 = SpatialUnitFactory.create(project=self.project)
        loc2 = SpatialUnitFactory.create(project=self.project)
        par = PartyFactory.create(project=self.project)
        rel = TenureRelationshipFactory.create(project=self.project)

        ContentObject.objects.create(resource=res, content_object=loc1)
        ContentObject.objects.create(resource=res, content_object=loc2)
        ContentObject.objects.create(resource=res, content_object=par)
        ContentObject.objects.create(resource=res, content_object=rel)

        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_basic.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test6-orig.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        # Rewrite the ES dump file to inject the resource ID
        with open(es_dump_path, 'r') as infile:
            es_dump_path = os.path.join(test_dir, 'test6.esjson')
            fwrite = open(es_dump_path, 'w')
            for line in infile:
                fwrite.write(line.replace('ID3', res.id))
            fwrite.close()

        exporter = AllExporter(self.project)
        zip_path, mime_type = exporter.make_download(es_dump_path)

        assert zip_path == os.path.join(test_dir, 'test6-res.zip')
        assert mime_type == ('application/zip')

        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 8
            assert 'data.xlsx' in files
            assert 'resources.xlsx' in files
            assert 'resources/baby_goat.jpeg' in files
            assert 'shape_files/README.txt' in files
            assert 'shape_files/point.dbf' in files
            assert 'shape_files/point.prj' in files
            assert 'shape_files/point.shp' in files
            assert 'shape_files/point.shx' in files

            myzip.extract('resources.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'resources.xlsx'))
            assert wb.get_sheet_names() == ['resources']

            ws = wb['resources']
            assert ws['A1'].value == 'id'
            assert ws['B1'].value == 'name'
            assert ws['C1'].value == 'description'
            assert ws['D1'].value == 'filename'
            assert ws['E1'].value == 'locations'
            assert ws['F1'].value == 'parties'
            assert ws['G1'].value == 'relationships'
            assert ws['A2'].value == res.id
            assert ws['B2'].value == "Goat"
            assert ws['C2'].value == "Let's pretend there's a description."
            assert ws['D2'].value == "baby_goat.jpeg"
            ids = ws['E2'].value.split(',')
            assert len(ids) == 2
            assert loc1.id in ids
            assert loc2.id in ids
            assert ws['F2'].value == par.id
            assert ws['G2'].value == rel.id

            myzip.extract('data.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'data.xlsx'))
            assert wb.get_sheet_names() == [
                'locations', 'parties', 'relationships']

            ws = wb['locations']
            assert ws['A1'].value == 'id'
            assert ws['B1'].value == 'geometry.ewkt'
            assert ws['C1'].value == 'type'
            assert ws['D1'].value == 'quality'
            assert ws['E1'].value == 'infrastructure'
            assert ws['A2'].value == 'ID0'
            assert ws['B2'].value == 'SRID=4326;POINT (1 1)'
            assert ws['C2'].value == 'PA'
            assert ws['D2'].value == 'point'
            assert ws['E2'].value == 'food, electricity'

            ws = wb['parties']
            assert ws['A1'].value == 'id'
            assert ws['B1'].value == 'name'
            assert ws['C1'].value == 'type'
            assert ws['D1'].value == 'notes'
            assert ws['E1'].value == 'gender'
            assert ws['F1'].value == 'homeowner'
            assert ws['G1'].value == 'dob'
            assert ws['H1'].value == 'number_of_members'
            assert ws['I1'].value == 'date_formed'
            assert ws['A2'].value == 'ID1'
            assert ws['B2'].value == "Cadastanaut"
            assert ws['C2'].value == 'IN'
            assert ws['D2'].value is None
            assert ws['E2'].value == 'm'
            assert ws['F2'].value == 'yes'
            assert ws['G2'].value == '1951-05-05'
            assert ws['H2'].value is None
            assert ws['I2'].value is None

            ws = wb['relationships']
            assert ws['A1'].value == 'id'
            assert ws['B1'].value == 'party_id'
            assert ws['C1'].value == 'spatial_unit_id'
            assert ws['D1'].value == 'tenure_type.id'
            assert ws['E1'].value == 'tenure_type.label'
            assert ws['F1'].value == 'notes'
            assert ws['A2'].value == 'ID2'
            assert ws['B2'].value == 'ID1'
            assert ws['C2'].value == 'ID0'
            assert ws['D2'].value == 'CU'
            assert ws['E2'].value == "Customary Rights"
            assert ws['F2'].value == "The best relationship!"

    def test_make_download_empty(self):
        ensure_dirs()
        original_es_dump_path = os.path.join(
            os.path.dirname(settings.BASE_DIR),
            'search/tests/files/test_es_dump_empty.esjson'
        )
        es_dump_path = os.path.join(test_dir, 'test7.esjson')
        shutil.copy(original_es_dump_path, es_dump_path)

        exporter = AllExporter(self.project)
        zip_path, mime_type = exporter.make_download(es_dump_path)

        assert zip_path == os.path.join(test_dir, 'test7-res.zip')
        assert mime_type == ('application/zip')

        with ZipFile(zip_path) as myzip:
            assert myzip.namelist() == ['data.xlsx']
            myzip.extract('data.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'data.xlsx'))
            sheetnames = wb.get_sheet_names()
            assert sheetnames == ['Sheet']
            assert wb['Sheet']['A1'].value is None
