import csv
import io
import json
import os
import pytest
import shutil

from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import PermissionDenied
from django.core.urlresolvers import reverse
from django.http import Http404
from django.template.loader import render_to_string
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.exceptions import PermissionDenied as APIPermissionDenied
from unittest.mock import patch
from urllib.parse import quote as url_quote
from zipfile import ZipFile

from tutelary.models import Policy, assign_user_policies
from skivvy import ViewTestCase, APITestCase

from accounts.tests.factories import UserFactory
from core.tests.utils.cases import UserTestCase
from core.tests.utils.files import make_dirs  # noqa
from organization.tests.factories import ProjectFactory
from spatial.models import SpatialUnit
from spatial.tests.factories import SpatialUnitFactory
from party.models import Party, TenureRelationship
from party.tests.factories import PartyFactory, TenureRelationshipFactory
from resources.models import Resource
from resources.tests.factories import ResourceFactory
from resources.tests.utils import clear_temp  # noqa
from resources.utils.io import ensure_dirs
from questionnaires.managers import create_attrs_schema
from questionnaires.tests import attr_schemas
from questionnaires.tests.factories import QuestionnaireFactory
from ..views import async
from .fake_results import get_fake_es_api_results


api_url = (
    settings.ES_SCHEME + '://' + settings.ES_HOST + ':' + settings.ES_PORT)
test_dir = os.path.join(settings.MEDIA_ROOT, 'temp')


def assign_policies(user):
    clauses = {
        'clause': [
            {
                'effect': 'allow',
                'object': ['project/*/*'],
                'action': ['project.view_private', 'project.download'],
            },
        ],
    }
    policy = Policy.objects.create(
        name='test-policy',
        body=json.dumps(clauses))
    assign_user_policies(user, policy)


class SearchAPITest(APITestCase, UserTestCase, TestCase):

    view_class = async.Search
    post_data = {
        'q': 'searching',
        'start': 10,
        'length': 20,
        'draw': 40,
    }

    def setup_models(self):
        self.user = UserFactory.create()
        assign_policies(self.user)
        self.project = ProjectFactory.create(slug='test-project')
        QuestionnaireFactory.create(project=self.project)

        content_type = ContentType.objects.get(
            app_label='party', model='party')
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.individual_party_xform_group,
            content_type=content_type,
        )
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.default_party_xform_group,
            content_type=content_type,
        )
        content_type = ContentType.objects.get(
            app_label='party', model='tenurerelationship')
        create_attrs_schema(
            project=self.project,
            dict=attr_schemas.tenure_relationship_xform_group,
            content_type=content_type,
        )

        self.su = SpatialUnitFactory.create(project=self.project, type='CB')
        self.party = PartyFactory.create(
            project=self.project,
            type='IN',
            attributes={
                'gender': 'm',
                'homeowner': 'yes',
                'dob': '1951-05-05'
            })
        self.tenure_rel = TenureRelationshipFactory.create(
            spatial_unit=self.su, party=self.party, project=self.project,
            attributes={'notes': 'PBS is the best.'})
        self.resource = ResourceFactory.create(project=self.project)

        self.results = get_fake_es_api_results(
            self.project, self.su, self.party, self.tenure_rel, self.resource)
        self.proj_result = self.results['hits']['hits'][0]
        self.su_result = self.results['hits']['hits'][1]
        self.party_result = self.results['hits']['hits'][2]
        self.tenure_rel_result = self.results['hits']['hits'][3]
        self.resource_result = self.results['hits']['hits'][4]

        self.query = 'searching'
        self.query_body = {
            'query': {
                'simple_query_string': {
                    'default_operator': 'and',
                    'query': self.query,
                }
            },
            'from': 10,
            'size': 20,
            'sort': {'_score': {'order': 'desc'}},
        }
        self.es_endpoint = '{}/project-{}/_search/'.format(api_url,
                                                           self.project.id)
        self.es_body = json.dumps(self.query_body, sort_keys=True)

    def setup_url_kwargs(self):
        return {
            'organization': self.project.organization.slug,
            'project': self.project.slug,
        }

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_results(self, mock_post, mock_get):
        mock_post.return_value.status_code = 200
        mock_post.return_value.json.return_value = {
            'hits': {
                'total': 100,
                'hits': [{
                    '_type': 'spatial',
                    '_source': {
                        'id': self.su.id,
                        'type': 'AP',
                        '@timestamp': 'TIMESTAMP',
                    },
                }],
            },
        }

        response = self.request(user=self.user, method='POST')
        expected_html = render_to_string(
            'search/search_result_item.html',
            context={'result': {
                'entity_type': self.su.ui_class_name,
                'url': self.su.get_absolute_url(),
                'main_label': "Apartment",
                'attributes': {},
            }}
        )
        assert response.status_code == 200
        assert response.content['data'] == [[expected_html]]
        assert response.content['recordsTotal'] == 100
        assert response.content['recordsFiltered'] == 100
        assert response.content['draw'] == 40
        assert response.content['timestamp'] == 'TIMESTAMP'
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_over_max_results(self, mock_post, mock_get):
        mock_post.return_value.status_code = 200
        mock_post.return_value.json.return_value = {
            'hits': {
                'total': settings.ES_MAX_RESULTS + 1000,
                'hits': [{
                    '_type': 'spatial',
                    '_source': {
                        'id': self.su.id,
                        'type': 'AP',
                        '@timestamp': 'TIMESTAMP',
                    },
                }],
            },
        }

        response = self.request(user=self.user, method='POST')
        expected_html = render_to_string(
            'search/search_result_item.html',
            context={'result': {
                'entity_type': self.su.ui_class_name,
                'url': self.su.get_absolute_url(),
                'main_label': "Apartment",
                'attributes': {},
            }}
        )
        assert response.status_code == 200
        assert response.content['data'] == [[expected_html]]
        assert response.content['recordsTotal'] == settings.ES_MAX_RESULTS
        assert response.content['recordsFiltered'] == settings.ES_MAX_RESULTS
        assert response.content['draw'] == 40
        assert response.content['timestamp'] == 'TIMESTAMP'
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_no_results(self, mock_post, mock_get):
        mock_post.return_value.status_code = 200
        mock_post.return_value.json.return_value = {
            'hits': {
                'total': 100,
                'hits': [],
            },
        }
        mock_get.return_value.status_code = 200
        mock_get.return_value.json.return_value = {
            'hits': {
                'hits': [{
                    '_source': {
                        '@timestamp': 'TIMESTAMP',
                    },
                }],
            },
        }

        response = self.request(user=self.user, method='POST')
        assert response.status_code == 200
        assert response.content['data'] == []
        assert response.content['recordsTotal'] == 100
        assert response.content['recordsFiltered'] == 100
        assert response.content['draw'] == 40
        assert response.content['timestamp'] == 'TIMESTAMP'
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )
        mock_get.assert_called_once_with(
            '{}/project-{}/project/_search/?q=*'.format(api_url,
                                                        self.project.id))

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_project_result(self, mock_post, mock_get):
        mock_post.return_value.status_code = 200
        mock_post.return_value.json.return_value = {
            'hits': {
                'total': 100,
                'hits': [{
                    '_type': 'project',
                    '_source': {
                        '@timestamp': 'TIMESTAMP',
                    },
                }],
            },
        }

        response = self.request(user=self.user, method='POST')
        assert response.status_code == 200
        assert response.content['data'] == []
        assert response.content['recordsTotal'] == 100
        assert response.content['recordsFiltered'] == 100
        assert response.content['draw'] == 40
        assert response.content['timestamp'] == 'TIMESTAMP'
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_null_id(self, mock_post, mock_get):
        mock_post.return_value.status_code = 200
        mock_post.return_value.json.return_value = {
            'hits': {
                'total': 100,
                'hits': [{
                    '_type': 'spatial',
                    '_source': {
                        'id': None,
                        '@timestamp': 'TIMESTAMP',
                    },
                }],
            },
        }

        response = self.request(user=self.user, method='POST')
        assert response.status_code == 200
        assert response.content['data'] == []
        assert response.content['recordsTotal'] == 100
        assert response.content['recordsFiltered'] == 100
        assert response.content['draw'] == 40
        assert response.content['timestamp'] == 'TIMESTAMP'
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_missing_query(self, mock_post, mock_get):
        response = self.request(
            user=self.user, method='POST', post_data={'q': None})
        assert response.status_code == 200
        assert response.content['data'] == []
        assert response.content['recordsTotal'] == 0
        assert response.content['recordsFiltered'] == 0
        assert response.content['draw'] == 40
        assert response.content['timestamp'] == ''
        mock_post.assert_not_called()
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_es_not_ok(self, mock_post, mock_get):
        response = self.request(user=self.user, method='POST')
        assert response.status_code == 200
        assert response.content['data'] == []
        assert response.content['recordsTotal'] == 0
        assert response.content['recordsFiltered'] == 0
        assert response.content['draw'] == 40
        assert response.content['error'] == 'unavailable'
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_nonexistent_org(self, mock_post, mock_get):
        response = self.request(user=self.user,
                                method='POST',
                                url_kwargs={'organization': 'evil-corp'})
        assert response.status_code == 404
        assert response.content['detail'] == "Project not found."
        mock_post.assert_not_called()
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_nonexistent_project(self, mock_post, mock_get):
        response = self.request(user=self.user,
                                method='POST',
                                url_kwargs={'project': 'world-domination'})
        assert response.status_code == 404
        assert response.content['detail'] == "Project not found."
        mock_post.assert_not_called()
        mock_get.assert_not_called()

    @patch('requests.get')
    @patch('requests.post')
    def test_post_with_unauthorized_user(self, mock_post, mock_get):
        response = self.request(method='POST')
        assert response.status_code == 403
        assert response.content['detail'] == APIPermissionDenied.default_detail
        mock_post.assert_not_called()
        mock_get.assert_not_called()

    @patch('requests.post')
    def test_query_es(self, mock_post):
        mock_post.return_value.status_code = 200
        mock_post.return_value.json.return_value = {
            'hits': {
                'total': 0,
                'hits': [],
            },
        }

        raw_results = self.view_class().query_es(
            self.project.id, self.query, 10, 20)
        assert raw_results == mock_post.return_value.json.return_value
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )

    @patch('requests.post')
    def test_query_es_not_ok(self, mock_post):
        mock_post.return_value.status_code = 404

        raw_results = self.view_class().query_es(
            self.project.id, self.query, 10, 20)
        assert raw_results is None
        mock_post.assert_called_once_with(
            self.es_endpoint,
            data=self.es_body,
            headers={'content-type': 'application/json'},
        )

    @patch('requests.get')
    def test_query_es_timestamp(self, mock_get):
        mock_get.return_value.status_code = 200
        mock_get.return_value.json.return_value = {
            'hits': {
                'hits': [{
                    '_source': {
                        '@timestamp': 'TIMESTAMP',
                    },
                }],
            },
        }

        timestamp = self.view_class().query_es_timestamp(self.project.id)
        assert timestamp == 'TIMESTAMP'
        mock_get.assert_called_once_with(
            '{}/project-{}/project/_search/?q=*'.format(api_url,
                                                        self.project.id))

    @patch('requests.get')
    def test_query_es_timestamp_not_ok(self, mock_get):
        mock_get.return_value.status_code = 404

        timestamp = self.view_class().query_es_timestamp(self.project.id)
        assert timestamp == "unknown"
        mock_get.assert_called_once_with(
            '{}/project-{}/project/_search/?q=*'.format(api_url,
                                                        self.project.id))

    def test_augment_result_location(self):
        augmented_result = self.view_class().augment_result(self.su_result)
        assert augmented_result['entity_type'] == "Location"
        assert augmented_result['url'] == self.su.get_absolute_url()
        assert augmented_result['main_label'] == "Apartment"
        assert augmented_result['attributes'] == []

    def test_augment_result_party(self):
        augmented_result = self.view_class().augment_result(self.party_result)
        assert augmented_result['entity_type'] == "Party"
        assert augmented_result['url'] == self.party.get_absolute_url()
        assert augmented_result['main_label'] == "Party in the USA"
        attributes = augmented_result['attributes']
        assert len(attributes) == 1
        assert ("Type", "Group") in attributes

    def test_augment_result_tenure_rel(self):
        augmented_result = self.view_class().augment_result(
            self.tenure_rel_result)
        assert augmented_result['entity_type'] == "Relationship"
        assert augmented_result['url'] == self.tenure_rel.get_absolute_url()
        assert augmented_result['main_label'] == "Customary Rights"
        attributes = augmented_result['attributes']
        assert len(attributes) == 2
        assert ("Party", "Party in the USA") in attributes
        assert ("Location type", self.su.name) in attributes

    def test_augment_result_resource(self):
        augmented_result = self.view_class().augment_result(
            self.resource_result)
        assert augmented_result['entity_type'] == "Resource"
        assert augmented_result['url'] == self.resource.get_absolute_url()
        assert augmented_result['main_label'] == "Goat"
        attributes = augmented_result['attributes']
        assert len(attributes) == 2
        assert ("Original file", 'baby_goat.jpeg') in attributes
        assert ("Description",
                "Let's pretend there's a description.") in attributes
        assert augmented_result['image'] == self.resource.thumbnail

    def test_augment_result_nonexistent_entity(self):
        assert self.view_class().augment_result({
            '_type': 'spatial',
            '_source': {},
        }) is None

    def test_get_entity_location(self):
        assert self.view_class().get_entity(
            'spatial', self.su_result['_source']) == self.su

    def test_get_entity_party(self):
        assert self.view_class().get_entity(
            'party', self.party_result['_source']) == self.party

    def test_get_entity_tenure_rel(self):
        assert self.view_class().get_entity(
            'party', self.tenure_rel_result['_source']) == self.tenure_rel

    def test_get_entity_resource(self):
        assert self.view_class().get_entity(
            'resource', self.resource_result['_source']) == self.resource

    def test_get_entity_null_id(self):
        assert self.view_class().get_entity('spatial', {'id': None}) is None

    def test_get_entity_nonexistent_id(self):
        assert self.view_class().get_entity('spatial', {'id': 'xx'}) is None

    def test_get_entity_unsupported_es_type(self):
        assert self.view_class().get_entity('project', {}) is None

    def test_get_main_label_location(self):
        assert self.view_class().get_main_label(
            SpatialUnit, self.su_result['_source']) == "Apartment"

    def test_get_main_label_party(self):
        assert self.view_class().get_main_label(
            Party, self.party_result['_source']) == "Party in the USA"

    def test_get_main_label_tenure_rel(self):
        assert self.view_class().get_main_label(
            TenureRelationship, self.tenure_rel_result['_source']
        ) == "Customary Rights"

    def test_get_main_label_resource(self):
        assert self.view_class().get_main_label(
            Resource, self.resource_result['_source']) == "Goat"

    def test_get_main_label_unsupported_model(self):
        assert self.view_class().get_main_label(
            type(self.project), {'name': "Project"}) == "Project"

    def test_get_main_label_invalid_location_type(self):
        assert self.view_class().get_main_label(
            SpatialUnit, {'type': 'XX'}) == "—"

    def test_get_main_label_invalid_tenure_rel_type(self):
        assert self.view_class().get_main_label(
            TenureRelationship, {'tenure_type_id': 'XX'}) == "—"

    def test_get_attributes_location(self):
        assert self.view_class().get_attributes(
            self.su, self.su_result['_source']) == []

    def test_get_attributes_party(self):
        attributes = self.view_class().get_attributes(
            self.party, self.party_result['_source'])
        assert len(attributes) == 1
        assert ("Type", "Group") in attributes

    def test_get_attributes_tenure_rel(self):
        attributes = self.view_class().get_attributes(
            self.tenure_rel, self.tenure_rel_result['_source'])
        assert len(attributes) == 2
        assert ("Party", "Party in the USA") in attributes
        assert ("Location type", self.su.name) in attributes

    def test_get_attributes_resource(self):
        attributes = self.view_class().get_attributes(
            self.resource, self.resource_result['_source'])
        assert len(attributes) == 2
        assert ("Original file", 'baby_goat.jpeg') in attributes
        assert ("Description",
                "Let's pretend there's a description.") in attributes

    def test_htmlize_result(self):
        augmented_result = {
            'entity_type': "Resource",
            'url': '/resource/detail/',
            'main_label': "Goat",
            'attributes': [
                ("Original file", 'baby_goat.jpeg'),
                ("Description", "Nothing to see here."),
            ],
            'image': settings.ICON_URL.format('mp3'),
        }
        expected = render_to_string('search/search_result_item.html',
                                    context={'result': augmented_result})
        assert self.view_class().htmlize_result(augmented_result) == expected


def mock_subprocess_run_curl(*args):
    assert args[0][0] == 'curl'
    ensure_dirs()
    original_es_dump_path = os.path.join(
        os.path.dirname(settings.BASE_DIR),
        'search/tests/files/test_es_dump_basic.esjson'
    )
    shutil.copy(original_es_dump_path, args[0][2])


@pytest.mark.usefixtures('clear_temp')
@pytest.mark.usefixtures('make_dirs')
class SearchExportAPITest(ViewTestCase, UserTestCase, TestCase):

    view_class = async.SearchExport
    post_data = {
        'q': 'searching',
        'type': 'all',
    }

    def setup_models(self):
        self.user = UserFactory.create()
        assign_policies(self.user)
        self.project = ProjectFactory.create(slug='test-project')

    def setup_url_kwargs(self):
        return {
            'organization': self.project.organization.slug,
            'project': self.project.slug,
        }

    # Note: The following 4 tests, which correspond to the 4 export formats,
    # skip django-skivvy in order to test that the downloaded file contents
    # was generated by the correct exporter class

    @patch('subprocess.run', new=mock_subprocess_run_curl)
    def test_post_shp_type(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('async:search:export', kwargs=self.setup_url_kwargs()),
            {'q': 'searching', 'type': 'shp'},
        )
        assert response.status_code == 200
        assert (response.__getitem__('content-disposition') ==
                'attachment; filename={}.zip'.format(self.project.slug))
        assert response.__getitem__('content-type') == 'application/zip'

        # Dump content to a file
        zip_path = os.path.join(test_dir, 'tmp.zip')
        f = open(zip_path, 'wb')
        f.write(response.content)
        f.close()

        # Sanity content checks (full checking is done in text_export.py)
        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 8
            assert 'README.txt' in files
            assert 'locations.csv' in files
            assert 'point.shp' in files

            with myzip.open('locations.csv') as csv_file:
                rows = list(csv.reader(io.TextIOWrapper(csv_file)))
                assert rows[0][0] == 'id'
                assert rows[1][0] == 'ID0'

    @patch('subprocess.run', new=mock_subprocess_run_curl)
    def test_post_xls_type(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('async:search:export', kwargs=self.setup_url_kwargs()),
            {'q': 'searching', 'type': 'xls'},
        )
        assert response.status_code == 200
        assert (response.__getitem__('content-disposition') ==
                'attachment; filename={}.xlsx'.format(self.project.slug))
        assert (response.__getitem__('content-type') ==
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet')

        # Dump content to a file
        xls_path = os.path.join(test_dir, 'tmp.xlsx')
        f = open(xls_path, 'wb')
        f.write(response.content)
        f.close()

        # Sanity content checks (full checking is done in text_export.py)
        wb = load_workbook(xls_path)
        assert wb.get_sheet_names() == [
            'locations', 'parties', 'relationships']
        assert wb['locations']['B1'].value == 'geometry.ewkt'
        assert wb['locations']['D1'].value is None
        assert wb['locations']['B2'].value == 'SRID=4326;POINT (1 1)'
        assert wb['locations']['D2'].value is None

    @patch('subprocess.run', new=mock_subprocess_run_curl)
    def test_post_res_type(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('async:search:export', kwargs=self.setup_url_kwargs()),
            {'q': 'searching', 'type': 'res'},
        )
        assert response.status_code == 200
        assert (response.__getitem__('content-disposition') ==
                'attachment; filename={}.zip'.format(self.project.slug))
        assert response.__getitem__('content-type') == 'application/zip'

        # Dump content to a file
        zip_path = os.path.join(test_dir, 'tmp.zip')
        f = open(zip_path, 'wb')
        f.write(response.content)
        f.close()

        # Sanity content checks (full checking is done in text_export.py)
        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 2
            assert 'resources.xlsx' in files
            assert 'resources/baby_goat.jpeg' in files

            myzip.extract('resources.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'resources.xlsx'))
            assert wb.get_sheet_names() == ['resources']
            assert wb['resources']['G1'].value == 'relationships'
            assert wb['resources']['A2'].value == 'ID3'
            assert wb['resources']['E2'].value is None
            assert wb['resources']['F2'].value is None
            assert wb['resources']['G2'].value is None

    @patch('subprocess.run', new=mock_subprocess_run_curl)
    def test_post_all_type(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('async:search:export', kwargs=self.setup_url_kwargs()),
            {'q': 'searching', 'type': 'all'},
        )
        assert response.status_code == 200
        assert (response.__getitem__('content-disposition') ==
                'attachment; filename={}.zip'.format(self.project.slug))
        assert response.__getitem__('content-type') == 'application/zip'

        # Dump content to a file
        zip_path = os.path.join(test_dir, 'tmp.zip')
        f = open(zip_path, 'wb')
        f.write(response.content)
        f.close()

        # Sanity content checks (full checking is done in text_export.py)
        with ZipFile(zip_path) as myzip:
            files = myzip.namelist()
            assert len(files) == 8
            assert 'data.xlsx' in files
            assert 'resources.xlsx' in files
            assert 'resources/baby_goat.jpeg' in files
            assert 'shape_files/README.txt' in files
            assert 'shape_files/point.shp' in files

            myzip.extract('resources.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'resources.xlsx'))
            assert wb.get_sheet_names() == ['resources']
            assert wb['resources']['G1'].value == 'relationships'
            assert wb['resources']['A2'].value == 'ID3'
            assert wb['resources']['E2'].value is None
            assert wb['resources']['F2'].value is None
            assert wb['resources']['G2'].value is None

            myzip.extract('data.xlsx', test_dir)
            wb = load_workbook(os.path.join(test_dir, 'data.xlsx'))
            assert wb.get_sheet_names() == [
                'locations', 'parties', 'relationships']
            assert wb['locations']['B1'].value == 'geometry.ewkt'
            assert wb['locations']['D1'].value is None
            assert wb['locations']['B2'].value == 'SRID=4326;POINT (1 1)'
            assert wb['locations']['D2'].value is None

    def test_post_with_missing_query(self):
        response = self.request(user=self.user,
                                method='POST',
                                post_data={'q': None})
        assert response.status_code == 400

    def test_post_with_invalid_type(self):
        response = self.request(user=self.user,
                                method='POST',
                                post_data={'type': 'nonsense'})
        assert response.status_code == 400

    def test_post_with_nonexistent_org(self):
        with pytest.raises(Http404):
            self.request(user=self.user,
                         method='POST',
                         url_kwargs={'organization': 'evil-corp'})

    def test_post_with_nonexistent_project(self):
        with pytest.raises(Http404):
            self.request(user=self.user,
                         method='POST',
                         url_kwargs={'project': 'world-domination'})

    def test_post_with_unauthorized_user(self):
        with pytest.raises(PermissionDenied):
            self.request(method='POST')

    @patch('subprocess.run')
    def test_query_es(self, mock_run):
        es_dump_path = self.view_class().query_es(
            'projectid', 'userid', 'query')
        expected_path_begin = os.path.join(
            settings.MEDIA_ROOT,
            'temp/projectid-userid-'
        )
        query_dsl = {
            'query': {
                'simple_query_string': {
                    'default_operator': 'and',
                    'query': 'query',
                }
            },
            'from': 0,
            'size': settings.ES_MAX_RESULTS,
            'sort': {'_score': {'order': 'desc'}},
        }
        query_dsl_param = url_quote(json.dumps(query_dsl), safe='')
        assert es_dump_path[0:len(expected_path_begin)] == expected_path_begin
        assert '.esjson' in es_dump_path
        mock_run.assert_called_once_with([
            'curl',
            '-o', es_dump_path,
            '-XGET',
            '{}/project-projectid/_data/?format=json&source={}'.format(
                api_url, query_dsl_param
            )
        ])
