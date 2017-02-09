import os
from collections import OrderedDict

from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from openpyxl import Workbook

from .base import Exporter

MIME_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class XLSExporter(Exporter):

    def write_items(self, worksheet, queryset, content_type, model_attrs):
        schema_attrs = self.get_schema_attrs(content_type)

        # build column labels
        attr_columns = OrderedDict()
        for a in model_attrs:
            attr_columns[a] = ''
        for _, attrs in schema_attrs.items():
            for a in attrs.values():
                if a.name not in attr_columns.keys():
                    attr_columns[a.name] = ''
        worksheet.append(list(attr_columns.keys()))

        # write data
        for i, item in enumerate(queryset):
            values = self.get_values(item, model_attrs, schema_attrs)
            data = attr_columns.copy()
            data.update(values)
            worksheet.append(list(data.values()))

    def write_locations(self):
        locations = self.project.spatial_units.all()
        if locations.count() == 0:
            return
        worksheet = self.workbook.create_sheet(title='locations')
        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        self.write_items(worksheet, locations, content_type,
                         ['id', 'geometry.ewkt', 'type'])

    def write_parties(self):
        parties = self.project.parties.all()
        if parties.count() == 0:
            return
        worksheet = self.workbook.create_sheet(title='parties')
        content_type = ContentType.objects.get(app_label='party',
                                               model='party')
        self.write_items(worksheet, parties, content_type,
                         ['id', 'name', 'type'])

    def write_relationships(self):
        relationships = self.project.tenure_relationships.all()
        if relationships.count() == 0:
            return
        worksheet = self.workbook.create_sheet(title='relationships')
        content_type = ContentType.objects.get(app_label='party',
                                               model='tenurerelationship')
        self.write_items(worksheet, relationships, content_type,
                         ['party_id', 'spatial_unit_id', 'tenure_type.id',
                          'tenure_type.label'])

    def make_download(self, f_name):
        path = os.path.join(settings.MEDIA_ROOT, 'temp/{}.xlsx'.format(f_name))
        self.workbook = Workbook(write_only=True)

        self.write_locations()
        self.write_parties()
        self.write_relationships()

        self.workbook.save(filename=path)
        return path, MIME_TYPE
