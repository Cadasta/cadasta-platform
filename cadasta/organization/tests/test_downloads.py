import csv
import os
import time
from zipfile import ZipFile

import pytest

from core.tests.utils.cases import UserTestCase
from core.tests.utils.files import make_dirs  # noqa
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.contrib.gis.geos import GEOSGeometry
from django.test import TestCase
from jsonattrs.models import Attribute, AttributeType, Schema
from openpyxl import Workbook, load_workbook
from organization.tests.factories import ProjectFactory
from party.tests.factories import PartyFactory, TenureRelationshipFactory
from resources.models import ContentObject
from resources.tests.factories import ResourceFactory
from resources.tests.utils import clear_temp  # noqa
from resources.utils.io import ensure_dirs
from spatial.tests.factories import SpatialUnitFactory
from spatial.models import SpatialUnit
from questionnaires.tests import factories as q_factories

from ..download.base import Exporter
from ..download.resources import ResourceExporter
from ..download.shape import ShapeExporter
from ..download.xls import XLSExporter


class BaseExporterTest(UserTestCase, TestCase):

    def test_init(self):
        project = ProjectFactory.build()
        exporter = Exporter(project)
        assert exporter.project == project

    def test_get_schema_attrs_empty(self):
        project = ProjectFactory.create()
        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        exporter = Exporter(project)
        assert exporter.get_schema_attrs(content_type) == {}

    def test_get_schema_attrs(self):
        project = ProjectFactory.create(current_questionnaire='123abc')
        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        schema = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc', ))
        text_type = AttributeType.objects.get(name='text')
        select_m_type = AttributeType.objects.get(name='select_multiple')
        Attribute.objects.create(
            schema=schema,
            name='key', long_name='Test field',
            attr_type=text_type, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=schema,
            name='key_2', long_name='Test field',
            attr_type=text_type, index=1,
            required=False, omit=True
        )
        Attribute.objects.create(
            schema=schema,
            name='key_3', long_name='Test select multiple field',
            attr_type=select_m_type, index=2,
            choices=['choice_1', 'choice_2', 'choice_3'],
            choice_labels=['Choice 1', 'Choice 2', 'Choice 3'],
            required=False, omit=False
        )

        exporter = Exporter(project)
        attrs = exporter.get_schema_attrs(content_type)
        assert len(attrs['DEFAULT']) == 2

    def test_get_values(self):
        project = ProjectFactory.create(current_questionnaire='123abc')
        questionnaire = q_factories.QuestionnaireFactory.create(id='123abc')
        question = q_factories.QuestionFactory.create(
            questionnaire=questionnaire,
            name='tenure_type',
            type='S1'
        )
        q_factories.QuestionOptionFactory.create(
            question=question,
            name='LH',
            label='Miete'
        )
        exporter = Exporter(project)
        content_type = ContentType.objects.get(app_label='party',
                                               model='tenurerelationship')
        schema = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc', ))
        text_type = AttributeType.objects.get(name='text')
        select_m_type = AttributeType.objects.get(name='select_multiple')
        Attribute.objects.create(
            schema=schema,
            name='key', long_name='Test field',
            attr_type=text_type, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=schema,
            name='key_2', long_name='Test select multiple field',
            attr_type=select_m_type, index=1,
            choices=['choice_1', 'choice_2', 'choice_3'],
            choice_labels=['Choice 1', 'Choice 2', 'Choice 3'],
            required=False, omit=False
        )

        item = TenureRelationshipFactory.create(project=project,
                                                tenure_type='LH',
                                                attributes={
                                                    'key': 'text',
                                                    'key_2': ['choice_1',
                                                              'choice_2']})
        model_attrs = ('id', 'party_id', 'spatial_unit_id', 'tenure_type',
                       'tenure_type_label')
        schema_attrs = exporter.get_schema_attrs(content_type)
        values = exporter.get_values(item, model_attrs, schema_attrs)
        assert values == {
            'id': item.id, 'party_id': item.party_id,
            'spatial_unit_id': item.spatial_unit_id,
            'tenure_type': 'LH',
            'tenure_type_label': 'Miete',
            'key': 'text', 'key_2': 'choice_1, choice_2',
        }

    def test_get_values_null_geom(self):
        project = ProjectFactory.create(current_questionnaire='123abc')
        exporter = Exporter(project)
        item = SpatialUnitFactory.create(
            project=project,
            geometry=None)
        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        model_attrs = ('id', 'geometry.wkt')
        schema_attrs = exporter.get_schema_attrs(content_type)
        values = exporter.get_values(item, model_attrs, schema_attrs)
        assert values == {'id': item.id, 'geometry.wkt': None}

    def test_get_values_with_conditional_selector(self):
        project = ProjectFactory.create(current_questionnaire='123abc')
        exporter = Exporter(project)
        content_type = ContentType.objects.get(app_label='party',
                                               model='party')
        schema = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc',))
        schema_in = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc', 'IN'))
        schema_gr = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc', 'GR'))
        text_type = AttributeType.objects.get(name='text')
        select_m_type = AttributeType.objects.get(name='select_multiple')
        Attribute.objects.create(
            schema=schema,
            name='key', long_name='Test field',
            attr_type=text_type, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=schema_in,
            name='key_2', long_name='Test select multiple field',
            attr_type=select_m_type, index=1,
            choices=['choice_1', 'choice_2', 'choice_3'],
            choice_labels=['Choice 1', 'Choice 2', 'Choice 3'],
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=schema_gr,
            name='gr_key', long_name='Test Group Field',
            attr_type=text_type, index=0,
            required=False, omit=False
        )

        # test individual attrs
        item = PartyFactory.create(project=project,
                                   name='Test Party',
                                   type='IN',
                                   attributes={
                                       'key': 'text',
                                       'key_2': ['choice_1',
                                                 'choice_2']})
        model_attrs = ('id', 'name', 'type')
        schema_attrs = exporter.get_schema_attrs(content_type)
        values = exporter.get_values(item, model_attrs, schema_attrs)
        assert values == {
            'id': item.id, 'name': item.name, 'type': item.type,
            'key': 'text', 'key_2': 'choice_1, choice_2'
        }

        # test group attrs
        item = PartyFactory.create(project=project,
                                   name='Test Party',
                                   type='GR',
                                   attributes={
                                       'key': 'text',
                                       'gr_key': 'Test Group Field'})
        model_attrs = ('id', 'name', 'type')
        schema_attrs = exporter.get_schema_attrs(content_type)
        values = exporter.get_values(item, model_attrs, schema_attrs)
        assert values == {
            'id': item.id, 'name': item.name, 'type': item.type,
            'key': 'text', 'gr_key': 'Test Group Field'
        }


@pytest.mark.usefixtures('clear_temp')
class ShapeTest(UserTestCase, TestCase):

    def setUp(self):
        super().setUp()
        self.project = ProjectFactory.create(current_questionnaire='123abc')
        questionnaire = q_factories.QuestionnaireFactory.create(id='123abc')
        question = q_factories.QuestionFactory.create(
            questionnaire=questionnaire,
            name='tenure_type',
            type='S1'
        )
        q_factories.QuestionOptionFactory.create(
            question=question,
            name='CR',
            label='Eigentum'
        )
        self.spatial_content_type = ContentType.objects.get(
            app_label='spatial', model='spatialunit'
        )
        sp_schema = Schema.objects.create(
            content_type=self.spatial_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', ))
        attr_type = AttributeType.objects.get(name='text')
        Attribute.objects.create(
            schema=sp_schema,
            name='key', long_name='Test field',
            attr_type=attr_type, index=0,
            required=False, omit=False
        )
        self.spatialunit_1 = SpatialUnitFactory.create(
            project=self.project,
            geometry='POINT (1 1)',
            attributes={'key': 'value 1'})
        self.spatialunit_2 = SpatialUnitFactory.create(
            project=self.project,
            geometry='POINT (2 2)',
            attributes={'key': 'value 2'})
        self.spatialunits = [self.spatialunit_1, self.spatialunit_2]
        self.spatial_attrs = ['id', 'geometry.ewkt']

        self.party_content_type = ContentType.objects.get(
            app_label='party', model='party'
        )
        pt_schema = Schema.objects.create(
            content_type=self.party_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', ))
        pt_schema_in = Schema.objects.create(
            content_type=self.party_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', 'IN'))
        pt_schema_gr = Schema.objects.create(
            content_type=self.party_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', 'GR'))
        attr_type_txt = AttributeType.objects.get(name='text')
        attr_type_int = AttributeType.objects.get(name='integer')
        attr_type_dec = AttributeType.objects.get(name='decimal')

        Attribute.objects.create(
            schema=pt_schema,
            name='default_attr', long_name='Test field',
            attr_type=attr_type_txt, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=pt_schema,
            name='default_int_attr', long_name='Test integer field',
            attr_type=attr_type_int, index=1,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=pt_schema_in,
            name='party_in', long_name='Test IN field',
            attr_type=attr_type_txt, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=pt_schema_gr,
            name='party_gr', long_name='Test GR field',
            attr_type=attr_type_txt, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=pt_schema_gr,
            name='party_gr_dec', long_name='Test GR dec field',
            attr_type=attr_type_dec, index=1,
            required=False, omit=False
        )
        party_in = PartyFactory.create(
            project=self.project, type='IN',
            attributes={'default_attr': 'Test Schema',
                        'default_int_attr': 1,
                        'party_in': 'Test IN attribute'}
        )
        party_gr = PartyFactory.create(
            project=self.project, type='GR',
            attributes={'default_attr': 'Another Test Schema',
                        'default_int_attr': 2,
                        'party_gr_dec': 2.333,
                        'party_gr': 'Test GR attribute'}
        )
        self.parties = [party_in, party_gr]

        self.tenurerelationship = TenureRelationshipFactory.create(
            project=self.project, party=self.parties[0],
            spatial_unit=self.spatialunit_1, tenure_type='CR'
        )

    def test_init(self):
        project = ProjectFactory.build()
        exporter = ShapeExporter(project)
        assert exporter.project == project

    def test_create_datasource(self):
        ensure_dirs()
        project = ProjectFactory.create()
        exporter = ShapeExporter(project)

        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/file')
        ds = exporter.create_datasource(dst_dir)
        assert (ds.GetName() ==
                os.path.join(settings.MEDIA_ROOT, 'temp/file'))
        ds.Destroy()

    def test_create_layer(self):
        ensure_dirs()
        project = ProjectFactory.create()
        exporter = ShapeExporter(project)

        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/file6')
        ds = exporter.create_datasource(dst_dir)
        for layer_type in [
                'point', 'linestring', 'polygon',
                'multilinestring', 'multipoint', 'multipolygon']:
            layer = exporter.create_layer(ds, layer_type)
            assert layer is not None
            assert layer.GetName() == layer_type
            assert layer.GetGeomType() in [1, 2, 3, 4]
            assert layer.GetLayerDefn().GetFieldCount() == 1
        ds.Destroy()

    def test_write_items(self):

        exporter = ShapeExporter(self.project)
        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/party')
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)
        filename = os.path.join(dst_dir, 'parties.csv')
        exporter.write_items(filename, self.parties, self.party_content_type,
                             ('id', 'name', 'type'))

        with open(filename) as csvfile:
            csvreader = csv.reader(csvfile)
            for i, row in enumerate(csvreader):
                assert len(row) == 8
                if i == 0:
                    assert row == [
                        'id', 'name', 'type', 'default_attr',
                        'default_int_attr', 'party_in', 'party_gr',
                        'party_gr_dec']
                if i == 1:
                    assert row == [
                        self.parties[0].id, self.parties[0].name,
                        self.parties[0].type, 'Test Schema', '1',
                        'Test IN attribute', '', ''
                    ]
                if i == 2:
                    assert row == [
                        self.parties[1].id, self.parties[1].name,
                        self.parties[1].type, 'Another Test Schema', '2',
                        '', 'Test GR attribute', '2.333'
                    ]
                if i > 2:
                    assert False, 'Too many rows in CSV'

    def test_write_features(self):
        ensure_dirs()
        project = ProjectFactory.create(current_questionnaire='123abc')
        exporter = ShapeExporter(project)

        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        schema = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc', ))
        attr_type = AttributeType.objects.get(name='text')
        Attribute.objects.create(
            schema=schema,
            name='geom_type', long_name='Test field',
            attr_type=attr_type, index=0,
            required=False, omit=False
        )

        su1 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;POINT (30 10)',
            attributes={'geom_type': 'point'})
        su2 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;LINESTRING (30 10, 10 30, 40 40)',
            attributes={'geom_type': 'linestring'})
        su3 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;POLYGON ((30 10, 40 40, 20 40, 10 20, 30 10))',
            attributes={'geom_type': 'polygon'})
        su4 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;'
                     'MULTIPOINT ((10 40), (40 30), (20 20), (30 10))',
            attributes={'geom_type': 'multipoint'})
        su5 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;'
                     'MULTILINESTRING ((10 10, 20 20, 10 40),'
                     '(40 40, 30 30, 40 20, 30 10))',
            attributes={'geom_type': 'multilinestring'})
        su6 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;'
                     'MULTIPOLYGON (((30 20, 45 40, 10 40, 30 20)),'
                     '((15 5, 40 10, 10 20, 5 10, 15 5)))',
            attributes={'geom_type': 'multipolygon'})
        su7 = SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;POLYGON EMPTY',
            attributes={'geom_type': 'empty'})
        su8 = SpatialUnitFactory.create(
            project=project,
            geometry=None,
            attributes={'geom_type': 'none'})

        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/file4')
        ds = exporter.create_datasource(dst_dir)
        filename = os.path.join(dst_dir, 'locations.csv')

        layers = exporter.write_features(ds, filename)
        assert len(layers.keys()) == 6
        for layer_name, layer in layers.items():
            su = SpatialUnit.objects.get(attributes={'geom_type': layer_name})
            geom = su.geometry
            feature = layer.GetNextFeature()
            assert geom.equals(GEOSGeometry(feature.geometry().ExportToWkt()))
            assert feature.GetFieldAsString('id') == su.id

            # Ensuring empty polygons are not added to shape
            assert su.id not in [su7.id, su8.id]
        ds.Destroy()

        with open(filename) as csvfile:
            csvreader = csv.reader(csvfile)
            for i, row in enumerate(csvreader):
                if i == 0:
                    head = ['id', 'type', 'area', 'geom_type']
                    assert row == head
                if i == 1:
                    assert row == [su1.id, su1.type, '0.0', 'point']
                if i == 2:
                    assert row == [su2.id, su2.type, '0.0', 'linestring']
                if i == 3:
                    assert row == [su3.id, su3.type, str(su3.area), 'polygon']
                if i == 4:
                    assert row == [su4.id, su4.type, '0.0', 'multipoint']
                if i == 5:
                    assert row == [su5.id, su5.type, '0.0', 'multilinestring']
                if i == 6:
                    assert row == [su6.id, su6.type, str(su6.area),
                                   'multipolygon']
                if i == 7:
                    assert row == [su7.id, su7.type, '0.0', 'empty']
                if i == 8:
                    assert row == [su8.id, su8.type, '0.0', 'none']

        # remove this so other tests pass
        os.remove(filename)

    def test_write_features_empty(self):
        project = ProjectFactory.create()
        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/file4')
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)
        filename = os.path.join(dst_dir, 'locations.csv')

        exporter = ShapeExporter(project)
        exporter.write_features((), filename)

        assert not os.path.exists(filename)

    def test_write_relationships_empty(self):
        project = ProjectFactory.create()
        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/rel')
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)
        filename = os.path.join(dst_dir, 'relationships.csv')

        exporter = ShapeExporter(project)
        exporter.write_relationships(filename)

        assert not os.path.exists(filename)

    def test_write_relationships(self):
        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/rels')
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)
        filename = os.path.join(dst_dir, 'releationships.csv')

        exporter = ShapeExporter(self.project)
        exporter.write_relationships(filename)

        with open(filename) as csvfile:
            csvreader = csv.reader(csvfile)
            for i, row in enumerate(csvreader):
                if i == 0:
                    assert row == ['id', 'party_id', 'spatial_unit_id',
                                   'tenure_type']
                else:
                    assert row == [
                        self.tenurerelationship.id, self.parties[0].id,
                        self.spatialunit_1.id, 'CR']

        # remove this so other tests pass
        os.remove(filename)

    def test_write_parties_empty(self):
        project = ProjectFactory.create()
        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/party')
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)
        filename = os.path.join(dst_dir, 'parties.csv')

        exporter = ShapeExporter(project)
        exporter.write_parties(filename)
        assert not os.path.exists(filename)

    def test_write_parties(self):
        dst_dir = os.path.join(settings.MEDIA_ROOT, 'temp/party')
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)
        filename = os.path.join(dst_dir, 'parties.csv')

        exporter = ShapeExporter(self.project)
        exporter.write_parties(filename)

        with open(filename) as csvfile:
            csvreader = csv.reader(csvfile)
            for i, row in enumerate(csvreader):
                if i == 0:
                    assert row == [
                        'id', 'name', 'type', 'default_attr',
                        'default_int_attr', 'party_in', 'party_gr',
                        'party_gr_dec'
                    ]
                if i == 1:
                    assert row == [
                        self.parties[0].id, self.parties[0].name, 'IN',
                        'Test Schema', '1', 'Test IN attribute', '', ''
                    ]
                if i == 2:
                    assert row == [
                        self.parties[1].id, self.parties[1].name,
                        self.parties[1].type, 'Another Test Schema', '2',
                        '', 'Test GR attribute', '2.333'
                    ]
                if i > 2:
                    assert False, 'Too many rows in CSV'

        # remove this so other tests pass
        os.remove(filename)

    def test_make_download(self):
        ensure_dirs()
        project = ProjectFactory.create(current_questionnaire='123abc')
        exporter = ShapeExporter(project)

        content_type = ContentType.objects.get(app_label='spatial',
                                               model='spatialunit')
        schema = Schema.objects.create(
            content_type=content_type,
            selectors=(project.organization.id, project.id, '123abc', ))
        attr_type = AttributeType.objects.get(name='text')
        Attribute.objects.create(
            schema=schema,
            name='key', long_name='Test field',
            attr_type=attr_type, index=0,
            required=False, omit=False
        )

        SpatialUnitFactory.create(
            project=project,
            geometry='POINT (1 1)',
            attributes={'key': 'value 1'})
        SpatialUnitFactory.create(
            project=project,
            geometry='SRID=4326;'
                     'MULTIPOLYGON (((30 20, 45 40, 10 40, 30 20)),'
                     '((15 5, 40 10, 10 20, 5 10, 15 5)))',
            attributes={'key': 'value 2'})

        path, mime = exporter.make_download('file5')

        assert path == os.path.join(settings.MEDIA_ROOT, 'temp/file5.zip')
        assert mime == 'application/zip'

        with ZipFile(path, 'r') as testzip:
            assert len(testzip.namelist()) == 10
            assert 'point.dbf' in testzip.namelist()
            assert 'point.prj' in testzip.namelist()
            assert 'point.shp' in testzip.namelist()
            assert 'point.shx' in testzip.namelist()
            assert 'multipolygon.dbf' in testzip.namelist()
            assert 'multipolygon.prj' in testzip.namelist()
            assert 'multipolygon.shp' in testzip.namelist()
            assert 'multipolygon.shx' in testzip.namelist()
            assert 'locations.csv' in testzip.namelist()
            assert 'README.txt' in testzip.namelist()


@pytest.mark.usefixtures('clear_temp')
class XLSTest(UserTestCase, TestCase):

    def setUp(self):
        super().setUp()
        self.project = ProjectFactory.create(current_questionnaire='123abc')
        questionnaire = q_factories.QuestionnaireFactory.create(id='123abc')
        question = q_factories.QuestionFactory.create(
            questionnaire=questionnaire,
            name='tenure_type',
            type='S1'
        )
        q_factories.QuestionOptionFactory.create(
            question=question,
            name='FREE',
            label='Eigentum'
        )
        self.spatial_content_type = ContentType.objects.get(
            app_label='spatial', model='spatialunit'
        )
        sp_schema = Schema.objects.create(
            content_type=self.spatial_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', ))
        attr_type_text = AttributeType.objects.get(name='text')
        attr_type_int = AttributeType.objects.get(name='integer')
        attr_type_dec = AttributeType.objects.get(name='decimal')
        Attribute.objects.create(
            schema=sp_schema,
            name='key', long_name='Test field',
            attr_type=attr_type_text, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=sp_schema,
            name='integer', long_name='Test integer field',
            attr_type=attr_type_int, index=1,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=sp_schema,
            name='decimal', long_name='Test decimal field',
            attr_type=attr_type_dec, index=2,
            required=False, omit=False
        )
        self.spatialunit_1 = SpatialUnitFactory.create(
            project=self.project,
            geometry='POINT (1 1)',
            attributes={'key': 'value 1', 'integer': '1', 'decimal': '1.0'})
        self.spatialunit_2 = SpatialUnitFactory.create(
            project=self.project,
            geometry='POINT (2 2)',
            attributes={'key': 'value 2', 'integer': '2', 'decimal': '2.0'})
        self.spatialunits = [self.spatialunit_1, self.spatialunit_2]
        self.spatial_attrs = ['id', 'geometry.ewkt']

        self.party_content_type = ContentType.objects.get(
            app_label='party', model='party'
        )
        pt_schema = Schema.objects.create(
            content_type=self.party_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', ))
        pt_schema_in = Schema.objects.create(
            content_type=self.party_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', 'IN'))
        pt_schema_gr = Schema.objects.create(
            content_type=self.party_content_type,
            selectors=(
                self.project.organization.id, self.project.id, '123abc', 'GR'))
        attr_type = AttributeType.objects.get(name='text')
        Attribute.objects.create(
            schema=pt_schema,
            name='default_attr', long_name='Test field',
            attr_type=attr_type, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=pt_schema_in,
            name='party_in', long_name='Test IN field',
            attr_type=attr_type, index=0,
            required=False, omit=False
        )
        Attribute.objects.create(
            schema=pt_schema_gr,
            name='party_gr', long_name='Test GR field',
            attr_type=attr_type, index=0,
            required=False, omit=False
        )
        party_in = PartyFactory.create(
            project=self.project, type='IN',
            attributes={'default_attr': 'Test Schema',
                        'party_in': 'Test IN attribute'}
        )
        party_gr = PartyFactory.create(
            project=self.project, type='GR',
            attributes={'default_attr': 'Another Test Schema',
                        'party_gr': 'Test GR attribute'}
        )
        self.parties = [party_in, party_gr]

        TenureRelationshipFactory.create(
            project=self.project, party=self.parties[0],
            spatial_unit=self.spatialunit_1,
            tenure_type='FREE'
        )

    def test_init(self):
        exporter = XLSExporter(self.project)
        assert exporter.project == self.project

    def test_write_items(self):
        workbook = Workbook()
        worksheet = workbook.create_sheet()
        worksheet.title = 'locations'

        exporter = XLSExporter(self.project)
        exporter.write_items(
            worksheet, self.spatialunits, self.spatial_content_type,
            self.spatial_attrs
        )

        assert worksheet['A1'].value == 'id'
        assert worksheet['B1'].value == 'geometry.ewkt'
        assert worksheet['C1'].value == 'key'
        assert worksheet['D1'].value == 'integer'
        assert worksheet['E1'].value == 'decimal'

        assert worksheet['A2'].value == self.spatialunit_1.id
        assert worksheet['B2'].value == self.spatialunit_1.geometry.ewkt
        assert worksheet['C2'].value == 'value 1'
        assert worksheet['D2'].value == '1'
        assert worksheet['E2'].value == '1.0'

        assert worksheet['A3'].value == self.spatialunit_2.id
        assert worksheet['B3'].value == self.spatialunit_2.geometry.ewkt
        assert worksheet['C3'].value == 'value 2'
        assert worksheet['D3'].value == '2'
        assert worksheet['E3'].value == '2.0'

    def test_write_locations(self):
        workbook = Workbook()
        exporter = XLSExporter(self.project)
        exporter.workbook = workbook
        exporter.write_locations()
        assert workbook['locations']
        assert len(list(workbook['locations'].rows)) == 3

    def test_write_locations_empty(self):
        workbook = Workbook()

        project = ProjectFactory.build()
        exporter = XLSExporter(project)
        exporter.workbook = workbook
        exporter.write_locations()
        with pytest.raises(KeyError) as e:
            workbook['locations']
        assert str(e.value) == "'Worksheet locations does not exist.'"

    def test_write_parties(self):
        workbook = Workbook()

        exporter = XLSExporter(self.project)
        exporter.workbook = workbook
        exporter.write_parties()
        assert workbook['parties']
        assert len(list(workbook['parties'].rows)) == 3
        assert workbook['parties']['C2'].value == 'IN'
        assert workbook['parties']['E2'].value == 'Test IN attribute'
        assert workbook['parties']['E3'].value == ''
        assert workbook['parties']['F2'].value == ''
        assert workbook['parties']['C3'].value == 'GR'
        assert workbook['parties']['F3'].value == 'Test GR attribute'
        assert workbook['parties']

    def test_write_parties_empty(self):
        workbook = Workbook()

        project = ProjectFactory.build()
        exporter = XLSExporter(project)
        exporter.workbook = workbook
        exporter.write_parties()
        with pytest.raises(KeyError) as e:
            workbook['parties']
        assert str(e.value) == "'Worksheet parties does not exist.'"

    def test_write_relationships(self):
        workbook = Workbook()

        exporter = XLSExporter(self.project)
        exporter.workbook = workbook

        exporter.write_relationships()
        assert workbook['relationships']
        assert len(list(workbook['relationships'].rows)) == 2

    def test_write_relationships_empty(self):
        workbook = Workbook()

        project = ProjectFactory.build()
        exporter = XLSExporter(project)
        exporter.workbook = workbook

        exporter.write_relationships()
        with pytest.raises(KeyError) as e:
            workbook['relationships']
        assert str(e.value) == "'Worksheet relationships does not exist.'"

    def test_make_download(self):
        ensure_dirs()
        exporter = XLSExporter(self.project)
        path, mime = exporter.make_download('file')

        assert path == os.path.join(settings.MEDIA_ROOT, 'temp/file.xlsx')
        assert (mime == 'application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheet')


@pytest.mark.usefixtures('clear_temp')
@pytest.mark.usefixtures('make_dirs')
class ResourcesTest(UserTestCase, TestCase):

    def test_init(self):
        project = ProjectFactory.build()
        exporter = ResourceExporter(project)
        assert exporter.project == project

    def test_make_resource_worksheet(self):
        ensure_dirs()
        project = ProjectFactory.create()
        exporter = ResourceExporter(project)

        data = [
            ['1', 'n1', 'd1', 'f1', 'l1', 'p1', 'r1'],
            ['2', 'n2', 'd2', 'f2', 'l2', 'p2', 'r2']
        ]

        res_path = exporter.make_resource_worksheet('test-res', data)
        wb = load_workbook(filename=res_path, read_only=True)
        sheet = wb[wb.get_sheet_names()[0]]

        expected_headers = ['id', 'name', 'description', 'filename',
                            'locations', 'parties', 'relationships']
        for i, h in enumerate(expected_headers):
            assert sheet[chr(i + 97) + '1'].value == h

        for i in range(0, len(data[0])):
            assert sheet[chr(i + 97) + '2'].value == data[0][i]
            assert sheet[chr(i + 97) + '3'].value == data[1][i]

    def test_pack_resource_data(self):
        project = ProjectFactory.create()
        exporter = ResourceExporter(project)

        res = ResourceFactory.create(project=project)
        loc = SpatialUnitFactory.create(project=project)
        loc2 = SpatialUnitFactory.create(project=project)
        par = PartyFactory.create(project=project)
        rel = TenureRelationshipFactory.create(project=project)

        ContentObject.objects.create(resource=res, content_object=loc)
        ContentObject.objects.create(resource=res, content_object=loc2)
        ContentObject.objects.create(resource=res, content_object=par)
        ContentObject.objects.create(resource=res, content_object=rel)

        packed = exporter.pack_resource_data(res)
        assert packed[0] == res.id
        assert packed[1] == res.name
        assert packed[2] == res.description
        assert packed[3] == res.original_file
        assert loc.id in packed[4]
        assert loc2.id in packed[4]
        assert par.id in packed[5]
        assert rel.id in packed[6]

    def test_make_download(self):
        ensure_dirs()
        project = ProjectFactory.create()
        exporter = ResourceExporter(project)

        ResourceFactory.create(project=project, original_file='res.png')
        ResourceFactory.create(project=project, original_file='res.png')
        ResourceFactory.create(project=project, original_file='resources.xlsx')
        deleted = ResourceFactory.create(
            project=project,
            original_file='image1.jpg',
            archived=True)

        t = round(time.time() * 1000)
        path, mime = exporter.make_download('res-test-' + str(t))
        assert path == os.path.join(settings.MEDIA_ROOT,
                                    'temp/res-test-{}.zip'.format(t))
        assert mime == 'application/zip'

        with ZipFile(path, 'r') as testzip:
            assert len(testzip.namelist()) == 4
            assert 'res.png' in testzip.namelist()
            assert 'res_1.png' in testzip.namelist()
            assert 'resources_1.xlsx' in testzip.namelist()
            assert 'resources.xlsx' in testzip.namelist()
            assert deleted.original_file not in testzip.namelist()
